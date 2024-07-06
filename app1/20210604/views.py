# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from requests_ntlm import HttpNtlmAuth
# from django.conf import settings

# Create your views here.

# from django.db import models
from datetime import timedelta
from .models import TestLog
# from .forms import UserForm,testplanForm
from rbac.models import UserInfo,Item_Spec
from rbac.service.init_permission import init_permission
# from xlwt import *
# from io import BytesIO
import datetime,re,os#,urllib,requests
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
# from django.utils.encoding import escape_uri_path
from django.http import HttpResponse
import json,pymysql,requests,time
from django.views.decorators.csrf import csrf_exempt
# from bs4 import BeautifulSoup
# from selenium import webdriver
# from urllib import parse

class workDays():
    def __init__(self, start_date, end_date, days_off=None):
        """days_off:休息日,默认周六日, 以0(星期一)开始,到6(星期天)结束, 传入tupple
        没有包含法定节假日,
        """
        self.start_date = start_date
        self.end_date = end_date
        self.days_off = days_off
        if self.start_date > self.end_date:
            self.start_date,self.end_date = self.end_date, self.start_date
        if days_off is None:
            self.days_off = 5,6
        # 每周工作日列表
        self.days_work = [x for x in range(7) if x not in self.days_off]

    def workDays(self):
        """实现工作日的 iter, 从start_date 到 end_date , 如果在工作日内,yield 日期
        """
        # 还没排除法定节假日
        tag_date = self.start_date
        while True:
            if tag_date > self.end_date:
                break
            if tag_date.weekday() in self.days_work:
                yield tag_date
            tag_date += datetime.timedelta(days=1)

    def daysCount(self):
        """工作日统计,返回数字"""
        return len(list(self.workDays()))

    def weeksCount(self, day_start=0):
        """统计所有跨越的周数,返回数字
        默认周从星期一开始计算
        """
        day_nextweek = self.start_date
        while True:
            if day_nextweek.weekday() == day_start:
                break
            day_nextweek += datetime.timedelta(days=1)
        # 区间在一周内
        if day_nextweek > self.end_date:
            return 1
        weeks = ((self.end_date - day_nextweek).days + 1)/7
        weeks = int(weeks)
        if ((self.end_date - day_nextweek).days + 1)%7:
            weeks += 1
        if self.start_date < day_nextweek:
            weeks += 1
        return weeks


@csrf_exempt
def judgebackmysql():
    databasename = "metestcourse"
    host_DB = '192.168.1.11'
    user_DB = 'edwin'
    passwd_DB = 'DCT@2019'

    try:
        db=pymysql.connect(host=host_DB, user=user_DB, passwd=passwd_DB, db=databasename, port=3306, charset="utf8mb4")
        cursor = db.cursor()
        judge = cursor.execute("show tables")
        # print(judge)
        return 0
    except:
        return ('Mysql backup server is offline,please contact administrator')

def judgetime(list):
    n=0
    i=0
    for i in list:
        n+=1
    while i<n:
        if list[i][1]>list[i+1][0]:
            del list[i]
            del list[i+1]
            new=[list[i][0],list[i+1][1]]
            list.insert(i,new)
            n-=1
        else:
            i += 1
    return list

@csrf_exempt
def login(request):
    # return HttpResponse("hello world")
    context = {}
    context['hello'] = '登錄'
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    # 2 form
    # if request.method == "POST":
    #
    #     login_form = UserForm(request.POST)
    #     message = "请检查填写的内容！"
    #     print (login_form.is_valid())
    #     if login_form.is_valid():
    #         Account = login_form.cleaned_data['Accountf']
    #         # print (Account)
    #         Password = login_form.cleaned_data['Passwordf']
    #         # print(Password)
    #         try:
    #             # print(Password)
    #             user = UserInfo.objects.get(Account=Account)
    #             # print (user.Password)
    #             if user.Password == Password:
    #                 return redirect('/index/')
    #             else:
    #                 message = "密码不正确！"
    #         except:
    #             message = "用户不存在！"
    #     return render(request, 'login/login.html', locals())
    #
    # login_form = UserForm()
    # return render(request, 'login/login.html', locals())
    # 1
    #     username = request.POST.get('username')
    #     password = request.POST.get('password')
    #
    #     if username and password:  # 确保用户名和密码都不为空
    #         username = username.strip()
    #         # 用户名字符合法性验证
    #         # 密码长度验证
    #         # 更多的其它验证.....
    #         try:
    #             user = UserInfo.objects.get(Account=username)
    #             print (user.Password)
    #             if user.Password == password:
    #                 return redirect('/index/')
    #             else:
    #                 message = "密码不正确！"
    #         except:
    #             message = "用户名不存在！"
    #         return render(request, 'login/login.html', {"message": message})
    # return render(request, 'login/login.html')
    # 3 session
    # 不允许重复登录
    if request.session.get('is_login', None):
        return redirect('/index/')
    # print(request.method)
    # print('test')

    if request.method == "POST":
        # login_form = UserForm(request.POST)
        message = "请检查填写的内容！"
        # if login_form.is_valid():
        Account = request.POST.get('inputEmail')
        Password = request.POST.get('inputPassword')
        user_obj = UserInfo.objects.filter(account=Account, password=Password).first()
        # print (Account)
        # print (Password)
        # print (user_obj)
        try:
            user = UserInfo.objects.get(account=Account)
            # print (user.password)
            if user.password == Password:
                # 往session字典内写入用户状态和数据,你完全可以往里面写任何数据，不仅仅限于用户相关！
                request.session['is_login'] = True
                request.session['user_id'] = user.id
                request.session['user_name'] = user_obj
                # request.session['Skin'] = "../static/src/blue.jpg"
                request.session.set_expiry(12*60*60)
                # print('11')
                # Skin=request.COOKIES.get('Skin')
                # if not Skin:
                #     Skin="../static/src/blue.jpg"
                # print(Skin)
                # print('21')
                init_permission(request, user_obj)  # 调用init_permission，初始化权限
                return redirect('/index/')
            else:
                message = "密码不正确！"
        except:
            message = "用户不存在！"
        return render(request, 'test.html', locals())


    return render(request, 'test.html', locals())

@csrf_exempt
def logout(request):
    # print('t')
    # print (request.session.get('is_login', None))
    if not request.session.get('is_login', None):
        # 如果本来就未登录，也就没有登出一说
        # print('logout')
        return redirect("/login/")
    #flush()方法是比较安全的一种做法，而且一次性将session中的所有内容全部清空，确保不留后患。但也有不好的地方，那就是如果你在session中夹带了一点‘私货’，会被一并删除，这一点一定要注意
    request.session.flush()
    # 或者使用下面的方法
    # del request.session['is_login']
    # del request.session['user_id']
    # del request.session['user_name']
    return redirect("/login/")

@csrf_exempt
def Change_Password(request):
    context = {}
    context['hello'] = 'ChangePassword'
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    if not request.session.get('is_login', None):
        # 如果本来就未登录，也就没有登出一说
        return redirect("/login/")
    # print (request.method)
    if request.method == "POST":
        OldPassword=request.POST.get('OldPassword')
        Password = request.POST.get('Password')
        Passwordc = request.POST.get('Confirm')
        user=request.session.get('user_name')
        userpass=UserInfo.objects.get(username=user).password
        # print(OldPassword,userpass)
        if OldPassword==userpass:
            if Password==Passwordc:
                # print(request.session.get('user_name', None))
                updatep = UserInfo.objects.filter(username=request.session.get('user_name', None))
                # print (updatep)
                # for e in updatep:
                #    print (e.password)
                updatep.update(password=Password)
                request.session.flush()
                return redirect("/login/")
            else:
                message="Password is not same"
                return render(request, 'changepassword.html', locals())
        else:
            message = "Incorrect Password"
            return render(request, 'changepassword.html', locals())
    return render(request, 'changepassword.html', locals())


@csrf_exempt
def Change_Skin(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    # print(request.method)
    Skin = request.COOKIES.get('Skin')
    # print(Skin)
    if not Skin:
        Skin = "../static/src/blue.jpg"
    # print(Skin)
    Render = render(request, 'ChangeSkin.html', locals())
    Redirect=redirect('/Change_Skin/')
    if request.method == "POST":
        if 'Change' in request.POST:
            Skinv = request.POST.get('Skin')

            # print (Skinv)

            # print('1')
            if Skinv=='blue':
                # request.session['Skin'] ="../static/src/blue.jpg"
                Redirect.set_cookie('Skin', "../static/src/blue.jpg",3600*24*30*12)
            if Skinv=='kiwi':
                # request.session['Skin'] ="../static/src/kiwi.jpg"
                Redirect.set_cookie('Skin', "../static/src/kiwi.jpg",3600*24*30*12)
            if Skinv=='sunny':
                # request.session['Skin'] ="../static/src/sunny.jpg"
                Redirect.set_cookie('Skin', "../static/src/sunny.jpg",3600*24*30*12)
            if Skinv=='yellow':
                # request.session['Skin'] ="../static/src/yellow.jpg"
                Redirect.set_cookie('Skin', "../static/src/yellow.jpg",3600*24*30*12)
            # print('2')
            return Redirect
            # return redirect('/index/')
    # return redirect('/index/')
    # print(Skin)
    return Render

@csrf_exempt
def index(request):
    # 跳轉頁面
    context = {}
    context['hello'] = '主界面'
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    if not request.session.get('is_login', None):
        return redirect('/login/')
    # print('2')
    # if request.session.get('is_login', None):
    #     return redirect('/index')

    # Test_form = testplanForm(request.POST)
    actstatus = ""
    actstatus_end = ""
    # Comments = ""
    message = ''
    messagec = ''
    messagep = ''
    messagei = ''
    message_se = ''
    End_time = "first"
    Result='default'
    Skin = request.COOKIES.get('Skin')
    # print(Skin)
    if not Skin:
        Skin = "../static/src/blue.jpg"
    # print(Skin)
    test_list = TestLog.objects.filter(Result="")
    # print('12')
    # print (request.method)
    # print (request.POST)
    # print (request.GET)

    if request.method == "POST":
        # print('11')
        # Test_form = testplanForm(request.POST)
        # Project_Unit = Test_form.cleaned_data['Unit']
        # Project = Project_Unit[:5]
        # Unit = Project_Unit[5:]
        # Phase = Test_form.cleaned_data['Phase']
        # Testitem = Test_form.cleaned_data['Item']
        # Comments = Test_form.cleaned_data['Comment']
        # dic = {'Project': Project, 'Unit': Unit, 'Phase': Phase, 'Testitem': Testitem}
        # test_log = TestLog()
        # update = TestLog.objects.filter(**dic)
        # print ('1')
        # if not update.End_time:
        #     actstatus = "disabled"
        #     print('2')
        # print (request.POST.get('Start'))
        # print(request.POST)
        # tt = TestLog.objects.all().values('Phase').distinct().order_by('Phase')
        # print (tt)#筛选去重排序
        if 'unitSearch' in request.POST:
            unit = 'SKUNO='+request.POST.get('units')
            # print(unit,'1')

            url = r'http://192.168.1.10/dct/api/ClientSvc/getSKUInfo'
            requests.adapters.DEFAULT_RETRIES = 5
            # s = requests.session()
            # s.keep_alive = False  # 关闭多余连接
            # getTestSpec=requests.get(url)
            headers = {'Connection': 'close'}
            try:
                r = requests.get(url, headers=headers)
                getTestSpec = requests.get(url, unit)
                # print (getTestSpec.url)
            except:
                time.sleep(1)


            targetURL = getTestSpec.url
            # url=r"http://127.0.0.1"

            url.split('\n')[0]
            # print url
            # 输入用户名和密码python requests实现windows身份验证登录

            try:
                getTestSpec = requests.get(targetURL, auth=HttpNtlmAuth('DCT\\administrator', 'DQA3`2018'))
            except:
                time.sleep(1)
                # print "try request agian"

            # print 1
            # print getTestSpec.url
            # print (getTestSpec.text)
            Customer=json.loads(getTestSpec.text)['Customer']
            Phase=json.loads(getTestSpec.text)['Phase']
            data={'Customer':Customer,'Phase':Phase}
            # print(data)
            if data['Phase']=='B(DVT)':
                data['Phase']='B(FVT)'
            if data['Phase']=='C(PVT)':
                data['Phase']='C(SIT)'
            # datas={}
            # print (data)

            return HttpResponse(json.dumps(data))
        if 'itemSearch' in request.POST:
            Testitem = request.POST.get('item')
            unit = request.POST.get('units')
            dic_Des = {'Item_I': Testitem}
            Item_Des = Item_Spec.objects.filter(**dic_Des)
            # print(dic_Des)
            # print(Item_Des)
            itemdes="Wrong Item"
            if Item_Des:
                for e in Item_Des:
                    # print (e)
                    itemdes = e.Item_Description
            # print(itemdes)
            return HttpResponse(itemdes)
        if 'Start' in request.POST:
            # print('4')
            # Test_form = testplanForm(request.POST)
            # if Test_form.is_valid():  # 必须要先验证否则提示object错误没有attribute 'cleaned_data'
            Customer = request.POST.get('Customer')
            if len(Customer) <= 10:
                Project_Unit = request.POST.get('Units')
                if len(Project_Unit)==10:
                    # print('t')
                    Project = Project_Unit[:5]
                    Unit = Project_Unit[5:]
                    # print(Project)
                    # print(Unit)
                    Phase = request.POST.get('Phase')
                    # print(Phase)
                    if len(Phase)<=10:
                        Testitem = request.POST.get('Item')
                        dic_i = {'Item_I': Testitem}
                        Item_spec = Item_Spec.objects.filter(**dic_i)
                        # print(Item_Des)
                        if Item_spec:
                            if request.POST.get('Comments'):
                                Comments = 'Start:'+request.POST.get('Comments')
                            else:
                                Comments=request.POST.get('Comments')
                            dic = {'Customer':Customer,'Project': Project, 'Unit': Unit, 'Phase': Phase,  'End_time': ""}
                            # print (dic)
                            check = TestLog.objects.filter(**dic)
                            dicr = {'Customer': Customer, 'Project': Project, 'Unit': Unit, 'Phase': Phase,
                                     'Result': ""}
                            update = TestLog.objects.filter(**dicr)
                            # print(check)
                            for e in check:
                                End_time = e.End_time
                                recordc_item=e.Testitem
                                # print(e.End_time)
                            # if End_time=="first":
                            #     test_log = TestLog()
                            #     test_log.Project = Project
                            #     test_log.Unit = Unit
                            #     test_log.Phase = Phase
                            #     test_log.Testitem = Testitem
                            #     test_log.Tester_id = request.session.get('user_name')
                            #     test_log.Comments = Comments
                            #     test_log.Start_time = datetime.datetime.now()
                            #     # test_log.End_time = End_time
                            #     test_log.save()
                            if not End_time:#有End_time为''
                                # actstatus = r'disabled="disabled"'
                                message_se =  r"%s %s %s %s is being occupied:%s no End_time&Result"%(Customer,Project,Phase,Unit,recordc_item)
                                # print(actstatus)
                            else:#check为空，End_time未定义
                                for e in update:
                                    Result = e.Result
                                    recordcu_item=e.Testitem
                                    # print (End_time)
                                if not Result:#有Result为''
                                    message_se = r"%s %s %s %s is being occupied:%s no Result,has End_time" % (Customer, Project, Phase, Unit, recordcu_item)
                                else:#update为空，Result未定义

                                    test_log = TestLog()
                                    test_log.Customer =Customer
                                    test_log.Project = Project.upper()
                                    test_log.Unit = Unit
                                    test_log.Phase = Phase
                                    test_log.Testitem = Testitem.upper()
                                    dic_i={'Item_I':Testitem}
                                    check_i = Item_Spec.objects.filter(**dic_i)
                                    for e in check_i:
                                        Item_Des=e.Item_Description
                                    # print(Item_Des)
                                    test_log.Item_Des = Item_Des
                                    test_log.Tester = request.session.get('user_name')
                                    test_log.Comments = Comments
                                    test_log.Start_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                    # test_log.End_time = End_time
                                    test_log.save()
                                    message_se = r"%s %s %s %s Start test %s"%(Customer,Project,Phase,Unit,Testitem)
                                    actstatus = ""
                            return render(request, 'index.html', locals())
                        else:
                            messagei = "Don't have this TestItem, Check again"
                            actstatus = ""
                            return render(request, 'index.html', locals())
                            # return HttpResponse(messagei)
                    else:
                        messagep="Please choose Phase"
                        actstatus = ""
                        return render(request, 'index.html', locals())
                else:
                    message="len of Units is not 10"
                    actstatus = ""
                    return render(request, 'index.html', locals())
            else:
                messagec="Please choose Customer"
                actstatus = ""
                return render(request, 'index.html', locals())

        if 'End' in request.POST:
            # print('3')
            # Test_form = testplanForm(request.POST)
            # if Test_form.is_valid():  # 必须要先验证否则提示object错误没有attribute 'cleaned_data'
            Customer = request.POST.get('Customer')
            if len(Customer) <= 10:
                Project_Unit = request.POST.get('Units')
                if len(Project_Unit) == 10:
                    Project = Project_Unit[:5]
                    Unit = Project_Unit[5:]
                    Phase = request.POST.get('Phase')
                    if len(Phase)<=10:
                        Testitem = request.POST.get('Item')
                        dic_i = {'Item_I': Testitem}
                        Item_spec = Item_Spec.objects.filter(**dic_i)
                        # print(Item_Des)
                        if Item_spec:
                            if request.POST.get('Comments'):
                                Comments = 'End:'+request.POST.get('Comments')
                            else:
                                Comments=request.POST.get('Comments')
                            dic = {'Customer':Customer,'Project': Project, 'Unit': Unit, 'Phase': Phase, 'End_time': ""}
                            check = TestLog.objects.filter(**dic)
                            dicr = {'Customer': Customer, 'Project': Project, 'Unit': Unit, 'Phase': Phase,
                                     'Result': ""}
                            update = TestLog.objects.filter(**dicr)
                            # print(update)
                            for e in check:
                                # print (e)
                                End_time = e.End_time
                                recordc_item = e.Testitem
                                # print (End_time)
                            if not End_time:#有End_time为''
                                # print('t')
                                if recordc_item.upper()==Testitem.upper():
                                    update.update(End_time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                                    for i in update:
                                        oldcomment = i.Comments
                                    if oldcomment:
                                        newcommnet = oldcomment+'\n'+Comments
                                    else:
                                        newcommnet = Comments
                                    update.update(Comments=newcommnet)
                                    message_se = r"%s %s %s %s %s test End" % (Customer, Project, Phase, Unit, Testitem)
                                    actstatus = ""
                                else:
                                    message_se=r"%s %s %s %s is being occupied:%s no Result&End_time"% (Customer, Project, Phase, Unit, recordc_item)
                                    # for e in update:
                                    #     e.End_time= datetime.datetime.now()
                            else:#check为空，End_time未定义
                                for e in update:
                                    Result = e.Result
                                    recordcu_item = e.Testitem
                                    # print (End_time)
                                if not Result:#有Result为''
                                    # actstatus_end = r'disabled="disabled"'
                                    message_se = r"%s %s %s %s is being occupied:%s no Result,has End_time"% (Customer, Project, Phase, Unit, recordcu_item)
                                    actstatus = ""
                                else:#update为空，Result未定义
                                    message_se = r"Please start test: %s %s %s %s %s first" % (
                                    Customer, Project, Phase, Unit, Testitem)
                                    actstatus = ""
                            return render(request, 'index.html', locals())
                        else:
                            messagei = "Don't have this TestItem, Check again"
                            actstatus = ""
                            return render(request, 'index.html', locals())
                    else:
                        messagep="Please choose Phase"
                        actstatus = ""
                        return render(request, 'index.html', locals())

                else:
                    message = "len of Units is not 10"
                    actstatus = ""
                    return render(request, 'index.html', locals())
            else:
                messagec = "Please choose Customer"
                actstatus = ""
                return render(request, 'index.html', locals())

        if 'PassButton' in request.POST:
            # print('3')
            # Test_form = testplanForm(request.POST)
            # if Test_form.is_valid():  # 必须要先验证否则提示object错误没有attribute 'cleaned_data'
            Customer = request.POST.get('Customer')
            if len(Customer) <= 10:
                Project_Unit = request.POST.get('Units')
                if len(Project_Unit) == 10:
                    Project = Project_Unit[:5]
                    Unit = Project_Unit[5:]
                    Phase = request.POST.get('Phase')
                    if len(Phase) <= 10:
                        Testitem = request.POST.get('item')
                        dic_i = {'Item_I': Testitem}
                        Item_spec = Item_Spec.objects.filter(**dic_i)
                        # print(Item_Des)
                        if Item_spec:
                            if request.POST.get('Comments'):
                                Comments = 'Result:'+request.POST.get('Comments')
                            else:
                                Comments=request.POST.get('Comments')

                            dice = {'Customer': Customer, 'Project': Project, 'Unit': Unit, 'Phase': Phase,
                                     'End_time': ""}
                            check = TestLog.objects.filter(**dice)
                            dicr = {'Customer': Customer, 'Project': Project, 'Unit': Unit, 'Phase': Phase,
                                    'Result': ""}
                            update = TestLog.objects.filter(**dicr)
                            # print(update,'1')
                            for e in check:
                                End_time = e.End_time
                                recordc_item = e.Testitem
                            if not End_time:#有End_time为''
                                message_se = r"%s %s %s %s is being occupied:%s no End_time&Result" % (Customer, Project, Phase, Unit, recordc_item)
                            else:#check为空，End_time未定义
                                for e in update:
                                    # print (e)
                                    Result = e.Result
                                    recordcu_item = e.Testitem
                                    # print (End_time,'2')
                                if not Result:#有Result为''
                                    if recordcu_item.upper() == Testitem.upper():
                                        # print('t')
                                        # print(update, '4')
                                        for i in update:
                                            oldcomment = i.Comments
                                            # print(oldcomment, '5')
                                        if oldcomment:
                                            newcommnet = oldcomment + '\n' + Comments
                                        else:
                                            newcommnet = Comments
                                        update.update(Comments=newcommnet)
                                        update.update(Result_time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                                        update.update(Result='Pass')
                                        message_se = r"%s %s %s %s %s test finished with Pass" % (Customer, Project, Phase, Unit, Testitem)
                                        actstatus = ""
                                        # for e in update:
                                        #     e.End_time= datetime.datetime.now()
                                    else:
                                        message_se = r"%s %s %s %s is being occupied:%s no Result has End_time" % (Customer, Project, Phase, Unit, recordcu_item)
                                else:#update为空，Result未定义
                                    # actstatus_end = r'disabled="disabled"'
                                    message_se = r"Please start the test:%s %s %s %s %s first"% (Customer, Project, Phase, Unit, Testitem)
                                    actstatus = ""
                                return render(request, 'index.html', locals())
                            return render(request, 'index.html', locals())
                        else:
                            messagei = "Don't have this TestItem, Check again"
                            actstatus = ""
                            return render(request, 'index.html', locals())
                    else:
                        messagep = "Please choose Phase"
                        actstatus = ""
                        return render(request, 'index.html', locals())

                else:
                    message = "len of Units is not 10"
                    actstatus = ""
                    return render(request, 'index.html', locals())
            else:
                messagec = "Please choose Customer"
                actstatus = ""
                return render(request, 'index.html', locals())
        if 'FailButton' in request.POST:
            # print('3')
            # Test_form = testplanForm(request.POST)
            # if Test_form.is_valid():  # 必须要先验证否则提示object错误没有attribute 'cleaned_data'
            Customer = request.POST.get('Customer')
            if len(Customer) <= 10:
                Project_Unit = request.POST.get('Units')
                if len(Project_Unit) == 10:
                    Project = Project_Unit[:5]
                    Unit = Project_Unit[5:]
                    Phase = request.POST.get('Phase')
                    if len(Phase) <= 10:
                        Testitem = request.POST.get('item')
                        dic_i = {'Item_I': Testitem}
                        Item_spec = Item_Spec.objects.filter(**dic_i)
                        # print(Item_Des)
                        if Item_spec:
                            if request.POST.get('Comments'):
                                Comments = 'Result:'+request.POST.get('Comments')
                            else:
                                Comments=request.POST.get('Comments')

                            dice = {'Customer': Customer, 'Project': Project, 'Unit': Unit, 'Phase': Phase,
                                     'End_time': ""}
                            check = TestLog.objects.filter(**dice)
                            dicr = {'Customer': Customer, 'Project': Project, 'Unit': Unit, 'Phase': Phase,
                                    'Result': ""}
                            update = TestLog.objects.filter(**dicr)
                            for e in check:
                                End_time = e.End_time
                                recordc_item = e.Testitem
                            if not End_time:#有End_time为''
                                message_se = r"%s %s %s %s is being occupied:%s no End_time&Result" % (Customer, Project, Phase, Unit, recordc_item)
                            else:#check为空，End_time未定义
                                for e in update:
                                    Result = e.Result
                                    recordcu_item = e.Testitem
                                    # print (End_time)
                                if not Result:#有Result为''
                                    if recordcu_item.upper() == Testitem.upper():
                                        # print('t')
                                        for i in update:
                                            oldcomment = i.Comments
                                        if oldcomment:
                                            newcommnet = oldcomment + '\n' + Comments
                                        else:
                                            newcommnet = Comments
                                        update.update(Comments=newcommnet)
                                        update.update(Result_time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                                        update.update(Result='Fail')
                                        message_se = r"%s %s %s %s %s test finished with Fail" % (Customer, Project, Phase, Unit, Testitem)
                                        actstatus = ""
                                        # for e in update:
                                        #     e.End_time= datetime.datetime.now()
                                    else:
                                        message_se = r"%s %s %s %s is being occupied:%s no Result has End_time" % (Customer, Project, Phase, Unit, recordcu_item)
                                else:#update为空，Result未定义
                                    # actstatus_end = r'disabled="disabled"'
                                    message_se = r"Please start the test:%s %s %s %s %s first"% (Customer, Project, Phase, Unit, Testitem)
                                    actstatus = ""
                                return render(request, 'index.html', locals())
                        else:
                            messagei = "Don't have this TestItem, Check again"
                            actstatus = ""
                            return render(request, 'index.html', locals())
                    else:
                        messagep = "Please choose Phase"
                        actstatus = ""
                        return render(request, 'index.html', locals())

                else:
                    message = "len of Units is not 10"
                    actstatus = ""
                    return render(request, 'index.html', locals())
            else:
                messagec = "Please choose Customer"
                actstatus = ""
                return render(request, 'index.html', locals())
    # print(message_se)
    # print(locals())
    return render(request, 'index.html', locals())

@csrf_exempt
def miniDashboardProject(request):
    # 跳轉頁面
    context = {}
    context['hello'] = '主界面'
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    if not request.session.get('is_login', None):
        return redirect('/login/')
    # print (request.method)
    # print(request.POST)
    Skin = request.COOKIES.get('Skin')
    if not Skin:
        Skin = "../static/src/blue.jpg"
    project_list = TestLog.objects.all().values('Project').distinct().order_by('Project')
    if request.method == "POST":
        if 'Search' in request.POST:
            JSONObject = {}

            Customer = request.POST.get("Customer")
            Project = request.POST.get("project")
            Phase = request.POST.get("Phase")
            # print (Project)
            if len(Customer) <= 10:
                if len(Project):
                    if len(Phase) <= 10:
                        request.session['Customer_Dash'] = Customer
                        request.session['Project_Dash'] = Project
                        request.session['Phase_Dash'] = Phase
                        dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
                        check_list = TestLog.objects.filter(**dic).values('Unit').distinct().order_by('Unit')
                        if check_list:
                            for i in check_list:
                                list = []
                                dici = {'Customer': Customer, 'Project': Project, 'Phase': Phase,'Unit':i['Unit']}
                                # test_list = TestLog.objects.filter(**dici).exclude(Result='').order_by('Start_time')
                                test_list = TestLog.objects.filter(**dici).order_by('Start_time')
                                # print (test_list)
                                for j in test_list:
                                    addlist = {}
                                    addlist['starttime']=j.Start_time
                                    addlist['endtime']=j.Result_time
                                    addlist['des'] = '%s:%s'%(j.Testitem,j.Item_Des)
                                    addlist['Result']=j.Result
                                    if j.Testitem=='DBR00000000':
                                        addlist['RD']=1
                                    else:
                                        addlist['RD'] = 0
                                    list.append(addlist)
                                # print(list)
                                # JSONObject[i['Testitem']]=judgetime(list)
                                JSONObject['a%s'%i['Unit']] = list
                            listtime = []
                            addtime = {}
                            check_starttime = TestLog.objects.filter(**dic).order_by('Start_time').first()
                            # print (check_starttime.Start_time)
                            start_time=str(check_starttime.Start_time)
                            addtime['starttime']='%s%s%s'%(start_time[0:4],start_time[5:7],start_time[8:10])

                            check_endtime = TestLog.objects.filter(**dic).order_by('Result_time').last()
                            dict = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Result_time': ''}
                            check_endtimenull = TestLog.objects.filter(**dict)
                            chechtime='default'
                            for e in check_endtimenull:
                                if not e.Result_time:
                                    chechtime = e.Result_time

                            if not chechtime:
                                addtime['endtime'] = datetime.datetime.now().strftime("%Y%m%d")
                            else:
                                end_time=str(check_endtime.Result_time)
                                addtime['endtime'] = '%s%s%s'%(end_time[0:4],end_time[5:7],end_time[8:10])
                            listtime.append(addtime)
                            JSONObject['time'] = listtime
                            # print (JSONObject)
                            return HttpResponse(json.dumps(JSONObject), content_type="application/json")

        if 'Waterfall' in request.POST:
            Customer = request.session.get('Customer_Dash', None)
            Project = request.session.get('Project_Dash', None)
            Phase = request.session.get('Phase_Dash', None)
            dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
            check_list = TestLog.objects.filter(**dic).values('Unit').distinct().order_by('Unit')
            if check_list:
                check_times = TestLog.objects.filter(**dic).values('Start_time').distinct().order_by('Start_time')
                check_timee = TestLog.objects.filter(**dic).values('Result_time').distinct().order_by('Result_time')
                starttime=check_times.first()['Start_time']
                checkresulttime=check_timee.first()['Result_time']
                resulttime=check_timee.last()['Result_time']
                # print (starttime)
                # print (check_times)
                # print(check_timee)
                # print (starttime)
                # print(checkresulttime)
                # print(resulttime)
                start = '%s-%s-%s'%(starttime[0:4],starttime[5:7],starttime[8:10])
                if checkresulttime:
                    end='%s-%s-%s'%(resulttime[0:4],resulttime[5:7],resulttime[8:10])
                else:
                    end=datetime.datetime.now().strftime("%Y-%m-%d")

                # '%s%s%s' % (end_time[0:4], end_time[5:7], end_time[8:10])
                # print (start,end)
                datestart = datetime.datetime.strptime(start, '%Y-%m-%d')
                startdate=datestart
                dateend = datetime.datetime.strptime(end, '%Y-%m-%d')
                # print (start,datestart)
                # print (end,dateend)
                wb = Workbook(write_only=True)
                excel_data = []
                h_data = ["Units No"]
                while datestart <= dateend:
                    h_data.append(datestart.strftime('%Y-%m-%d'))
                    datestart += datetime.timedelta(days=1)
                    # print (datestart.strftime('%Y-%m-%d'))
                # print(h_data)
                # work = workDays(startdate, dateend)
                # for i in work.workDays():  # 获取每一个工作日期
                #     print(i)
                n=0
                for d in h_data:
                    # print(d,h_data[n])
                    if d == 'Units No':
                        n+=1
                        continue
                    flag = 0
                    work = workDays(startdate, dateend)
                    for i in work.workDays():  # 获取每一个工作日期
                        # print(d,str(i))
                        if re.match(d, str(i)):
                            flag = 1
                            break
                    # print(flag)
                    if not flag:
                        h_data[n]+='双休'
                        # print(d)
                    n+=1
                # print(h_data)
                excel_data.append(h_data)
                for i in check_list:
                    data = ['%s %s %s %s' % (Customer, Project, Phase, i['Unit'])]
                    dici = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Unit': i['Unit']}
                    # test_list = TestLog.objects.filter(**dici).exclude(Result='').order_by('Start_time')
                    test_list = TestLog.objects.filter(**dici).order_by('Start_time')
                    edata={}
                    for item in h_data:
                        edata['item']=''
                        if item=='Units No':
                            continue
                        for j in test_list:
                            unit_starttime=j.Start_time
                            unit_resultteime= j.Result_time
                            # print (unit_starttime)
                            # print(unit_resultteime)
                            unit_start='%s-%s-%s' % (unit_starttime[0:4], unit_starttime[5:7], unit_starttime[8:10])
                            unit_result = '%s-%s-%s' % (unit_resultteime[0:4], unit_resultteime[5:7], unit_resultteime[8:10])
                            if not unit_resultteime:
                                if unit_start <= item:
                                    if edata['item']:
                                        edata['item']=edata['item']+'/'+j.Testitem
                                    else:
                                        edata['item']+=j.Testitem
                            else:
                                if unit_start<=item<=unit_result:
                                    # print ('yes')
                                    if edata['item']:
                                        edata['item']=edata['item']+'/'+j.Testitem
                                    else:
                                        edata['item']+=j.Testitem
                        data.append(edata['item'])
                    excel_data.append(data)
                ws = wb.create_sheet()
                for line in excel_data:
                    ws.append(line)
                response = HttpResponse(save_virtual_workbook(wb), content_type='application/octet-stream')
                response['Content-Disposition'] = 'attachment; filename=Waterfall.xls'
                return response
        if 'jian' in request.POST:
            Customer = request.session.get('Customer_Dash', None)
            Project = request.session.get('Project_Dash', None)
            Phase = request.session.get('Phase_Dash', None)
            dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
            check_list = TestLog.objects.filter(**dic).values('Unit').distinct().order_by('Unit')
            if check_list:
                check_times = TestLog.objects.filter(**dic).values('Start_time').distinct().order_by('Start_time')
                check_timee = TestLog.objects.filter(**dic).values('Result_time').distinct().order_by('Result_time')
                starttime = check_times.first()['Start_time']
                checkresulttime = check_timee.first()['Result_time']
                resulttime = check_timee.last()['Result_time']
                # print (starttime)
                # print (check_times)
                # print(check_timee)
                # print (starttime)
                # print(checkresulttime)
                # print(resulttime)
                start = '%s-%s-%s' % (starttime[0:4], starttime[5:7], starttime[8:10])
                if checkresulttime:
                    end = '%s-%s-%s' % (resulttime[0:4], resulttime[5:7], resulttime[8:10])
                else:
                    end = datetime.datetime.now().strftime("%Y-%m-%d")

                # '%s%s%s' % (end_time[0:4], end_time[5:7], end_time[8:10])
                # print (start,end)
                datestart = datetime.datetime.strptime(start, '%Y-%m-%d')
                startdate = datestart
                dateend = datetime.datetime.strptime(end, '%Y-%m-%d')
                # print (start,datestart)
                # print (end,dateend)
                wb = Workbook(write_only=True)
                excel_data = []
                h_data = ["Units No"]
                while datestart <= dateend:
                    h_data.append(datestart.strftime('%Y-%m-%d'))
                    datestart += datetime.timedelta(days=1)
                    # print (datestart.strftime('%Y-%m-%d'))
                # print(h_data)
                # work = workDays(startdate, dateend)
                # for i in work.workDays():  # 获取每一个工作日期
                #     print(i)
                n = 0
                for d in h_data:
                    # print(d,h_data[n])
                    if d == 'Units No':
                        n += 1
                        continue
                    flag = 0
                    work = workDays(startdate, dateend)
                    for i in work.workDays():  # 获取每一个工作日期
                        # print(d,str(i))
                        if re.match(d, str(i)):
                            flag = 1
                            break
                    # print(flag)
                    if not flag:
                        h_data[n] += '双休'
                        # print(d)
                    n += 1
                # print(h_data)
                excel_data.append(h_data)
                for i in check_list:
                    data = ['%s %s %s %s' % (Customer, Project, Phase, i['Unit'])]
                    dici = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Unit': i['Unit']}
                    # test_list = TestLog.objects.filter(**dici).exclude(Result='').order_by('Start_time')
                    test_list = TestLog.objects.filter(**dici).order_by('Start_time')
                    edata = {}
                    for item in h_data:
                        edata['item'] = 0
                        if item == 'Units No':
                            continue
                        for j in test_list:
                            unit_starttime = j.Start_time
                            unit_resultteime = j.Result_time
                            # print (unit_starttime)
                            # print(unit_resultteime)
                            unit_start = '%s-%s-%s' % (unit_starttime[0:4], unit_starttime[5:7], unit_starttime[8:10])
                            unit_result = '%s-%s-%s' % (
                            unit_resultteime[0:4], unit_resultteime[5:7], unit_resultteime[8:10])
                            if not unit_resultteime:
                                if unit_start <= item:

                                    edata['item'] += 1
                            else:
                                if unit_start <= item <= unit_result:

                                    edata['item'] += 1
                        if edata['item'] == 0:
                            edata['item'] = ''
                        data.append(edata['item'])
                    excel_data.append(data)
                ws = wb.create_sheet()
                for line in excel_data:
                    ws.append(line)
                response = HttpResponse(save_virtual_workbook(wb), content_type='application/octet-stream')
                response['Content-Disposition'] = 'attachment; filename=Waterfall.xls'
                return response

    return render(request, 'mini-project.html', locals())

@csrf_exempt
def DashboardProject(request):
    # 跳轉頁面
    context = {}
    context['hello'] = '主界面'
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    if not request.session.get('is_login', None):
        return redirect('/login/')
    # print (request.method)
    # print(request.POST)
    Skin = request.COOKIES.get('Skin')
    if not Skin:
        Skin = "../static/src/blue.jpg"
    project_list = TestLog.objects.all().values('Project').distinct().order_by('Project')
    if request.method == "POST":
        if 'Search' in request.POST:
            JSONObject = {}

            Customer = request.POST.get("Customer")
            Project = request.POST.get("project")
            Phase = request.POST.get("Phase")
            # print(len(Customer), Customer)
            # print (Project)
            if len(Customer) <= 10:
                if len(Project):
                    if len(Phase) <= 10:
                        request.session['Customer_Dash'] = Customer
                        request.session['Project_Dash'] = Project
                        request.session['Phase_Dash'] = Phase
                        dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
                        check_list = TestLog.objects.filter(**dic).values('Unit').distinct().order_by('Unit')
                        if check_list:
                            for i in check_list:
                                list = []
                                dici = {'Customer': Customer, 'Project': Project, 'Phase': Phase,'Unit':i['Unit']}
                                # test_list = TestLog.objects.filter(**dici).exclude(Result='').order_by('Start_time')
                                test_list = TestLog.objects.filter(**dici).order_by('Start_time')
                                # print (test_list)
                                for j in test_list:
                                    addlist = {}
                                    addlist['starttime']=j.Start_time
                                    addlist['endtime']=j.Result_time
                                    addlist['des'] = '%s:%s'%(j.Testitem,j.Item_Des)
                                    addlist['Result']=j.Result
                                    if j.Testitem=='DBR00000000' or j.Testitem=='BFT00000000':
                                        addlist['RD']=1
                                    else:
                                        addlist['RD'] = 0
                                    list.append(addlist)
                                # print(list)
                                # JSONObject[i['Testitem']]=judgetime(list)
                                JSONObject['a%s'%i['Unit']] = list
                            listtime = []
                            addtime = {}
                            check_starttime = TestLog.objects.filter(**dic).order_by('Start_time').first()
                            # print (check_starttime.Start_time)
                            start_time=str(check_starttime.Start_time)
                            addtime['starttime']='%s%s%s'%(start_time[0:4],start_time[5:7],start_time[8:10])

                            check_endtime = TestLog.objects.filter(**dic).order_by('Result_time').last()
                            dict = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Result_time': ''}
                            check_endtimenull = TestLog.objects.filter(**dict)
                            chechtime='default'
                            for e in check_endtimenull:
                                if not e.Result_time:
                                    chechtime = e.Result_time

                            if not chechtime:
                                addtime['endtime'] = datetime.datetime.now().strftime("%Y%m%d")
                            else:
                                end_time=str(check_endtime.Result_time)
                                addtime['endtime'] = '%s%s%s'%(end_time[0:4],end_time[5:7],end_time[8:10])
                            listtime.append(addtime)
                            JSONObject['time'] = listtime
                            # print (JSONObject)
                            return HttpResponse(json.dumps(JSONObject), content_type="application/json")
        if 'Export' in request.POST:
            Customer = request.session.get('Customer_Dash', None)
            Project = request.session.get('Project_Dash', None)
            Phase = request.session.get('Phase_Dash', None)
            dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
            check_list = TestLog.objects.filter(**dic).values('Unit').distinct().order_by('Unit')

            # print (check_list)
            num = []
            if check_list:
                for i in check_list:
                    n = 0
                    dici = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Unit': i['Unit']}
                    # test_list = TestLog.objects.filter(**dici).exclude(Result='').order_by('Start_time')
                    test_list = TestLog.objects.filter(**dici).order_by('Start_time')
                    for j in test_list:
                        n += 1
                    num.append(n)
                num.sort(reverse=True)  # 从大小到

                wb = Workbook(write_only=True)
                m = 1
                excel_data = []
                h_data = ["Units No"]
                while m <= num[0]:
                    h_data.append('Record %s' % m)
                    m += 1
                # print(h_data)
                excel_data.append(h_data)
                for i in check_list:
                    data = ['%s %s %s %s' % (Customer, Project, Phase, i['Unit'])]
                    dici = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Unit': i['Unit']}
                    # test_list = TestLog.objects.filter(**dici).exclude(Result='').order_by('Start_time')
                    test_list = TestLog.objects.filter(**dici).order_by('Start_time')
                    for j in test_list:
#                         data.append('''Start_time:         %s
# End_time:           %s
# Result_time:       %s
# %s:  %s
# Result:              %s by %s Commorent:%s''' % (
#                             j.Start_time, j.End_time, j.Result_time, j.Testitem, j.Item_Des, j.Result,
#                             j.Tester, j.Comments))
                            data.append('Start_time: %s\n' % j.Start_time+'End_time: %s\n'%j.End_time+'Result_time: %s\n'%j.Result_time+'%s:  %s\n'%(j.Testitem, j.Item_Des)+'Result: %s by %s Commorent:%s'%(j.Result,
    j.Tester, j.Comments))
                    # print(data)
                    excel_data.append(data)
                # print(excel_data)
                ws = wb.create_sheet()
                for line in excel_data:
                    ws.append(line)
                response = HttpResponse(save_virtual_workbook(wb), content_type='application/octet-stream')
                response['Content-Disposition'] = 'attachment; filename=Unitslog.xls'
                return response

    return render(request, 'DashboardProject.html', locals())

@csrf_exempt
def DashboardUnits(request):
    # 跳轉頁面
    context = {}
    JSONObject={}
    context['hello'] = '主界面'
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    # print(request.method)
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin')
    if not Skin:
        Skin = "../static/src/blue.jpg"
    # print(request.is_ajax())
    # if request.is_ajax():
    # defaultkey1=TestLog.objects.all().order_by('Start_time').last()
    # defaultkey2=str(defaultkey1).split('+')
    # # print(defaultkey2)
    # Customer=defaultkey2[1]
    # Project = defaultkey2[2]
    # Phase = defaultkey2[4]
    # dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
    # # print (dic)
    # checkx = TestLog.objects.filter(**dic).values('Unit').distinct().order_by('Unit')
    # checky = TestLog.objects.filter(**dic).values('Testitem').distinct().order_by('Testitem')
    #
    # l = 0
    # for i in checkx:
    #     l += 1
    #     keyi = 'x0y%s' % l
    #     # print(i)
    #     JSONObject[keyi] = i['Unit']
    #     h = 0
    #     for j in checky:
    #         h += 1
    #         if l == 1:
    #             keyj = 'x%sy0' % h
    #             # print(i)
    #             JSONObject[keyj] = j['Testitem']
    #
    #         num = 0
    #         dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Unit': i['Unit'],
    #                'Testitem': j['Testitem']}
    #         check_num = TestLog.objects.filter(**dic)
    #         # print ('tttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt')
    #         for e in check_num:
    #             if e.End_time:
    #                 # print(e)
    #                 num += 1
    #         key = 'x%sy%s' % (h, l)
    #         if num:
    #             JSONObject[key] = str(num)
    # if h < 13:
    #     h = 13
    # else:
    #     h += 1
    # if l < 10:
    #     l = 10
    # else:
    #     l += 1
    # JSONObject['x'] = str(h + 1)
    # JSONObject['y'] = str(l + 1)
    project_list = TestLog.objects.all().values('Project').distinct().order_by('Project')
    if request.method == "POST":
        JSONObject={}

        Customer = request.POST.get("Customer")
        Project = request.POST.get("project")
        Phase = request.POST.get("Phase")
        # JSONObject['x1y0'] = Project
        # print (Customer,Project,Phase)
        # print (type(json.dumps(JSONObject)))
        # print (type(JSONObject))
        # print(request.is_ajax())
        if len(Customer) <= 10:
            if len(Project):
                if len(Phase) <= 10:
                    dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
                    # print (dic)
                    checkx = TestLog.objects.filter(**dic).values('Unit').distinct().order_by('Unit')
                    checky = TestLog.objects.filter(**dic).values('Testitem').distinct().order_by('Testitem')

                    l=1
                    num_units=0
                    if checkx:
                        for i in checkx:
                            l += 1
                            num_units+=1
                            keyi = 'x0y%s' % l
                            # print(i)
                            JSONObject[keyi] = i['Unit']
                            h = 0

                            for j in checky:
                                item_units_num = 0
                                h += 1
                                dic_i = {'Customer': Customer, 'Project': Project, 'Phase': Phase,
                                       'Testitem': j['Testitem']}

                                if l==2:
                                    # item_units = TestLog.objects.filter(**dic_i).values('Unit').distinct()
                                    item_units = TestLog.objects.filter(**dic_i)
                                    # print(item_units)
                                    keyj = 'x%sy0' % h
                                    keyj_sumary = 'x%sy1' % h
                                    if item_units:
                                        for e in item_units:
                                            item_units_num+=1

                                    # print(i)
                                    dic_Des={'Item_I':j['Testitem']}
                                    Item_Des=Item_Spec.objects.filter(**dic_Des)
                                    # print(Item_Des)
                                    if Item_Des:
                                        for e in Item_Des:
                                            itemdes=e.Item_Description
                                            spec=e.Sample_Demand
                                            # print (itemdes)
                                    else:
                                        itemdes=''
                                    JSONObject[keyj] ='''%s%s,
                                    %s''' %(j['Testitem'],itemdes,spec)
                                    # JSONObject[keyj_sumary]='%sunits'%item_units_num
                                    JSONObject[keyj_sumary] = '%s' % item_units_num

                                num=0
                                dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase, 'Unit': i['Unit'],
                                       'Testitem': j['Testitem']}
                                check_num = TestLog.objects.filter(**dic)
                                # print ('tttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt')
                                for e in check_num:
                                    # if e.End_time:
                                        # print(e)
                                        num+=1
                                key='x%sy%s'%(h,l)
                                if num:
                                    JSONObject[key]=str(num)
                        if h<13:
                            h=13
                        else:
                            h+=1
                        if l<10:
                            l=10
                        else:
                            l+=1
                        JSONObject['x'] = str(h+1)
                        JSONObject['y'] = str(l+1)
                        JSONObject['x0y1']="""%sunits"""% num_units
                        #
                        # print(h,l,JSONObject)
                        # print (checkx)
                        # print(checky)
                        return HttpResponse(json.dumps(JSONObject), content_type="application/json")


        # return render(request, 'DashboardUnits.html', {'JSONObject':json.dumps(JSONObject)})

    # return HttpResponse(json.dumps(JSONObject), content_type="application/json")
    return render(request, 'DashboardUnits.html',locals())

@csrf_exempt
def SearchExport(request):
    # 跳轉頁面

    context = {}
    JSONObject={}
    context['hello'] = '主界面'
    # print(request.method)
    # global exportdefault
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    if not request.session.get('is_login', None):
        return redirect('/login/')
    # print (request.method)
    Skin = request.COOKIES.get('Skin')
    if not Skin:
        Skin = "../static/src/blue.jpg"
    Customer = request.GET.get('Customer')
    Project = request.GET.get('Project')
    Phase = request.GET.get('Phase')
    dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
    # print (dic)
    test_list = TestLog.objects.filter(**dic)
    project_list=TestLog.objects.all().values('Project').distinct().order_by('Project')
    if request.method == "POST":
        if 'Search' in request.POST:
            Customer = request.POST.get('Customer')
            Project = request.POST.get('Project')
            Phase = request.POST.get('Phase')
            # print (Project)
            # JSONObject['x1y0'] = Project
            # print (Customer,Project,Phase)
            # print (type(json.dumps(JSONObject)))
            # print (type(JSONObject))
            # print(request.is_ajax())
            if len(Customer) <= 10:
                if len(Project):
                    if len(Phase) <= 10:
                        dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
                        # print (dic)
                        test_list = TestLog.objects.filter(**dic)
                        request.session['Customer'] = Customer
                        request.session['Project'] = Project
                        request.session['Phase'] = Phase

                        # print (type(test_list))
                        return render(request, 'SearchExport.html', locals())
        if 'Export' in request.POST:
            Customer=request.session.get('Customer', None)
            Project = request.session.get('Project', None)
            Phase = request.session.get('Phase', None)
            dic = {'Customer': Customer, 'Project': Project, 'Phase': Phase}
            test_list = TestLog.objects.filter(**dic)
            # for i in test_list:
            #     print(i)
            if test_list:
                #创建工作薄
                # ws = Workbook(encoding='utf-8')
                # w = ws.add_sheet(u"Test log")
                # w.write(0, 0, "Customer")
                # w.write(0, 1, u"Project")
                # w.write(0, 2, u"Unit")
                # w.write(0, 3, u"Phase")
                # w.write(0, 4, u"Testitem")
                # w.write(0, 5, u"Tester")
                # w.write(0, 6, u"Comments")
                # w.write(0, 7, u"Start_time")
                # w.write(0, 8, u"End_time")
                # # 写入数据
                # excel_row = 1
                # for obj in test_list:
                #     data_Customer= obj.Customer
                #     data_Project = obj.Project
                #     data_Unit = obj.Unit
                #     data_Phase = obj.Phase
                #     dada_Testitem = obj.Testitem
                #     # dada_Tester= obj.Tester
                #     dada_Comments = obj.Comments
                #     dada_Start_time = obj.Start_time
                #     dada_End_time = obj.End_time
                #     w.write(excel_row, 0, data_Customer)
                #     w.write(excel_row, 1, data_Project)
                #     w.write(excel_row, 2, data_Unit)
                #     w.write(excel_row, 3, data_Phase)
                #     w.write(excel_row, 4, dada_Testitem)
                #     # w.write(excel_row, 5, dada_Tester)
                #     w.write(excel_row, 6, dada_Comments)
                #     w.write(excel_row, 7, dada_Start_time)
                #     w.write(excel_row, 8, dada_End_time)
                #     excel_row += 1
                # # 检测文件是够存在
                # # 方框中代码是保存本地文件使用，如不需要请删除该代码
                # ###########################
                # # exist_file = os.path.exists("test.xls")
                # # if exist_file:
                # #     os.remove(r"test.xls")
                # # ws.save("test.xls")
                # ############################
                # sio = BytesIO()
                # ws.save(sio)
                # sio.seek(0)
                # response = HttpResponse(sio.getvalue(), content_type='application/vnd.ms-excel')
                # response['Content-Disposition'] = 'attachment; filename=test.xls'
                # response.write(sio.getvalue())
                # return response
                # file_name = 'Testlog.txt'
                # content = ...
                # response = HttpResponse(content, content_type='application/octet-stream')
                # response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path(file_name))
                # return response
                wb = Workbook(write_only=True)
                excel_data = [["Customer", 'Project',"Phase", 'Unit',"Tester","Testitem",'Result',"Start_time","End_time",'Result_time',"Comments",'Item_Des']]
                for obj in test_list:
                    data=[obj.Customer,obj.Project,obj.Phase,obj.Unit,obj.Tester,obj.Testitem,obj.Result,obj.Start_time,obj.End_time,obj.Result_time,obj.Comments,obj.Item_Des]
                    excel_data.append(data)
                # print (excel_data)
                ws = wb.create_sheet()
                for line in excel_data:
                    ws.append(line)
                response = HttpResponse(save_virtual_workbook(wb), content_type='application/octet-stream')
                response['Content-Disposition'] = 'attachment; filename=Testlog.xls'
                return response

            return render(request, 'SearchExport.html', locals())


        # return render(request, 'DashboardUnits.html', {'JSONObject':json.dumps(JSONObject)})

    return render(request, 'SearchExport.html', locals())

@csrf_exempt
def Manage(request):
    # 跳轉頁面
    context = {}
    context['hello'] = '主界面'
    # if judgebackmysql():
    #     return HttpResponse(judgebackmysql())
    return redirect('/admin/')